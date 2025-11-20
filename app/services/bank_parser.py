"""
Bank Statement Parser Service
Supports CSV, CAMT.053 (ISO 20022 XML), MT940 (SWIFT), and PDF formats
"""

import csv
import hashlib
import logging
import re
import zipfile
from decimal import Decimal, InvalidOperation
from datetime import datetime
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

# Third-party parsers
import mt940
from pycamt.parser import Camt053Parser
from defusedxml import ElementTree as ET

try:
    import pdfplumber
except ImportError:  # pragma: no cover - optional dependency
    pdfplumber = None

from app.models.bank_transactions import (
    BankTransaction,
    BankStatement,
    BankStatementFormat,
    TransactionType,
    TransactionStatus,
    MatchStatus,
)

logger = logging.getLogger(__name__)


class BankStatementParser:
    """
    Universal bank statement parser supporting multiple formats
    """

    def __init__(self, organization_id: str, bank_account_id: str):
        self.organization_id = organization_id
        self.bank_account_id = bank_account_id

    def parse_file(
        self,
        file_content: bytes,
        file_name: str,
        format_type: Optional[BankStatementFormat],
        imported_by: Optional[str] = None,
    ) -> Tuple[BankStatement, List[BankTransaction]]:
        """
        Parse bank statement file and return statement + transactions.
        Automatically detects the format when not provided (or when the
        provided format fails) so that CSV, CAMT.053, MT940, and PDF files can
        be uploaded without manual selection.

        Args:
            file_content: Raw file bytes
            file_name: Original filename
            format_type: Statement format (CSV, CAMT053, MT940)
            imported_by: User ID who imported

        Returns:
            Tuple of (BankStatement, List[BankTransaction])
        """
        # Calculate file hash to prevent duplicates
        file_hash = hashlib.sha256(file_content).hexdigest()

        # Build ordered list of formats to attempt (user-provided first, detected second)
        formats_to_try: List[BankStatementFormat] = []
        if format_type:
            formats_to_try.append(format_type)

        detected_format = self.detect_format(file_content, file_name)
        if detected_format not in formats_to_try:
            formats_to_try.append(detected_format)

        if (
            self._looks_like_pdf(file_content, file_name)
            and BankStatementFormat.PDF not in formats_to_try
        ):
            formats_to_try.append(BankStatementFormat.PDF)

        errors: List[str] = []

        for fmt in formats_to_try:
            try:
                if fmt == BankStatementFormat.CSV:
                    return self._parse_csv(file_content, file_name, file_hash, imported_by)
                if fmt == BankStatementFormat.CAMT053:
                    return self._parse_camt053(file_content, file_name, file_hash, imported_by)
                if fmt == BankStatementFormat.MT940:
                    return self._parse_mt940(file_content, file_name, file_hash, imported_by)
                if fmt == BankStatementFormat.PDF:
                    return self._parse_pdf(file_content, file_name, file_hash, imported_by)
            except Exception as exc:
                logger.warning(
                    "Parsing as %s failed: %s. Will attempt fallback formats if available.",
                    fmt.value,
                    exc,
                )
                errors.append(f"{fmt.value}: {exc}")
                continue

        attempted = ", ".join(fmt.value for fmt in formats_to_try)
        detail = "; ".join(errors) if errors else "No parser attempts were executed."
        raise ValueError(
            f"Failed to parse bank statement. Tried formats: {attempted}. Details: {detail}"
        )

    def _parse_csv(
        self,
        file_content: bytes,
        file_name: str,
        file_hash: str,
        imported_by: Optional[str],
    ) -> Tuple[BankStatement, List[BankTransaction]]:
        """
        Parse CSV bank statement

        Expected CSV columns:
        - date, value_date, description, amount, currency, balance, reference,
          counterparty_name, counterparty_account
        """
        try:
            content_str = self._get_csv_string(file_content, file_name)
            # Normalize line endings and handle quoted fields properly
            content_str = content_str.replace('\r\n', '\n').replace('\r', '\n')
            content_str = self._strip_leading_metadata_lines(content_str)

            # Create CSV reader with proper configuration to handle various CSV formats
            csv_reader = csv.DictReader(
                StringIO(content_str),
                skipinitialspace=True,
                quoting=csv.QUOTE_MINIMAL
            )

            transactions = []
            opening_balance = None
            closing_balance = None
            total_debits = 0.0
            total_credits = 0.0
            min_date = None
            max_date = None
            row_errors: List[str] = []

            for row_num, row in enumerate(csv_reader, start=1):
                try:
                    normalized_row = self._normalize_row_keys(row)

                    # Parse transaction date
                    trans_date = self._parse_date(
                        self._get_first_value(normalized_row, "date", "transaction_date")
                    )
                    value_date = (
                        self._parse_date(self._get_first_value(normalized_row, "value_date"))
                        or trans_date
                    )

                    # Update date range
                    if min_date is None or trans_date < min_date:
                        min_date = trans_date
                    if max_date is None or trans_date > max_date:
                        max_date = trans_date

                    # Parse amount
                    amount, trans_type = self._extract_amount_and_type(normalized_row)

                    # Update totals
                    if trans_type == TransactionType.DEBIT:
                        total_debits += amount
                    else:
                        total_credits += amount

                    # Parse balance
                    balance = self._get_first_value(normalized_row, "balance", "balance_after")
                    if balance:
                        balance_value = self._parse_decimal(balance)
                        if opening_balance is None:
                            opening_balance = balance_value
                        closing_balance = balance_value

                    # Create transaction
                    transaction = BankTransaction(
                        organization_id=self.organization_id,
                        bank_account_id=self.bank_account_id,
                        transaction_date=trans_date,
                        value_date=value_date,
                        transaction_type=trans_type,
                        amount=amount,
                        currency=self._get_first_value(normalized_row, "currency") or "EUR",
                        reference=self._get_first_value(normalized_row, "reference"),
                        description=self._get_first_value(normalized_row, "description", "details"),
                        counterparty_name=self._get_first_value(normalized_row, "counterparty_name", "counterparty"),
                        counterparty_account=self._get_first_value(normalized_row, "counterparty_account", "counterparty_iban"),
                        balance_after=self._parse_decimal(balance) if balance else None,
                        status=TransactionStatus.PENDING,
                        match_status=MatchStatus.UNMATCHED,
                        imported_by=imported_by,
                        raw_data=self._sanitize_raw_row(row),
                    )

                    transactions.append(transaction)

                except Exception as e:
                    error_message = f"Row {row_num}: {e}"
                    row_errors.append(error_message)
                    logger.error(f"Error parsing CSV row {row_num}: {e}")
                    continue

            if not transactions:
                sample_errors = "; ".join(row_errors[:3]) if row_errors else "No parsable rows found."
                raise ValueError(f"No valid transactions were parsed from CSV. {sample_errors}")

            if row_errors:
                logger.warning(
                    "CSV parsing skipped %s rows due to validation errors. Sample: %s",
                    len(row_errors),
                    "; ".join(row_errors[:3]),
                )

            # Create statement
            statement = BankStatement(
                organization_id=self.organization_id,
                bank_account_id=self.bank_account_id,
                format=BankStatementFormat.CSV,
                file_name=file_name,
                file_hash=file_hash,
                statement_date=datetime.utcnow(),
                from_date=min_date or datetime.utcnow(),
                to_date=max_date or datetime.utcnow(),
                opening_balance=opening_balance or 0.0,
                closing_balance=closing_balance or 0.0,
                total_debits=total_debits,
                total_credits=total_credits,
                transaction_count=len(transactions),
                imported_by=imported_by,
            )

            return statement, transactions

        except Exception as e:
            logger.error(f"Error parsing CSV file: {e}")
            raise ValueError(f"Failed to parse CSV file: {str(e)}")

    def _parse_camt053(
        self,
        file_content: bytes,
        file_name: str,
        file_hash: str,
        imported_by: Optional[str],
    ) -> Tuple[BankStatement, List[BankTransaction]]:
        """
        Parse CAMT.053 (ISO 20022 XML) bank statement

        CAMT.053 is the modern XML-based standard for bank statements
        """
        try:
            # Parse XML using pycamt library
            xml_string = file_content.decode("utf-8")
            parser = Camt053Parser(xml_string)

            # Get statement info
            stmt_info = parser.get_statement_info()
            transactions_data = parser.get_transactions()

            transactions = []
            total_debits = 0.0
            total_credits = 0.0

            # Parse transactions
            for trans_data in transactions_data:
                try:
                    # Determine transaction type
                    credit_debit = trans_data.get("credit_debit_indicator", "")
                    is_credit = credit_debit == "CRDT"
                    trans_type = (
                        TransactionType.CREDIT
                        if is_credit
                        else TransactionType.DEBIT
                    )
                    amount = abs(float(trans_data.get("amount", 0)))

                    # Update totals
                    if trans_type == TransactionType.DEBIT:
                        total_debits += amount
                    else:
                        total_credits += amount

                    # Parse dates
                    booking_date = trans_data.get("booking_date")
                    value_date = trans_data.get("value_date")

                    if isinstance(booking_date, str):
                        booking_date = self._parse_date(booking_date)
                    if isinstance(value_date, str):
                        value_date = self._parse_date(value_date)

                    # Create transaction
                    transaction = BankTransaction(
                        organization_id=self.organization_id,
                        bank_account_id=self.bank_account_id,
                        transaction_date=booking_date or value_date or datetime.utcnow(),
                        value_date=value_date or booking_date or datetime.utcnow(),
                        booking_date=booking_date,
                        transaction_type=trans_type,
                        amount=amount,
                        currency=trans_data.get("currency", "EUR"),
                        transaction_id=trans_data.get("entry_reference"),
                        reference=trans_data.get("remittance_information"),
                        counterparty_name=trans_data.get("counterparty_name"),
                        counterparty_iban=trans_data.get("counterparty_account"),
                        description=trans_data.get("remittance_information"),
                        status=TransactionStatus.PENDING,
                        match_status=MatchStatus.UNMATCHED,
                        imported_by=imported_by,
                        raw_data=trans_data,
                    )

                    transactions.append(transaction)

                except Exception as e:
                    logger.error(f"Error parsing CAMT transaction: {e}")
                    continue

            # Parse statement balances
            opening_balance = 0.0
            closing_balance = 0.0
            from_date = datetime.utcnow()
            to_date = datetime.utcnow()

            if stmt_info:
                opening_balance = float(stmt_info.get("opening_balance", {}).get("amount", 0))
                closing_balance = float(stmt_info.get("closing_balance", {}).get("amount", 0))

                from_date_str = stmt_info.get("from_date")
                to_date_str = stmt_info.get("to_date")

                if from_date_str:
                    from_date = self._parse_date(from_date_str)
                if to_date_str:
                    to_date = self._parse_date(to_date_str)

            # Create statement
            statement = BankStatement(
                organization_id=self.organization_id,
                bank_account_id=self.bank_account_id,
                format=BankStatementFormat.CAMT053,
                file_name=file_name,
                file_hash=file_hash,
                statement_number=stmt_info.get("statement_id") if stmt_info else None,
                statement_date=datetime.utcnow(),
                from_date=from_date,
                to_date=to_date,
                opening_balance=opening_balance,
                closing_balance=closing_balance,
                total_debits=total_debits,
                total_credits=total_credits,
                transaction_count=len(transactions),
                imported_by=imported_by,
            )

            return statement, transactions

        except Exception as e:
            logger.error(f"Error parsing CAMT.053 file: {e}")
            raise ValueError(f"Failed to parse CAMT.053 file: {str(e)}")

    def _parse_mt940(
        self,
        file_content: bytes,
        file_name: str,
        file_hash: str,
        imported_by: Optional[str],
    ) -> Tuple[BankStatement, List[BankTransaction]]:
        """
        Parse MT940 (SWIFT) bank statement

        MT940 is the legacy text-based SWIFT standard
        """
        try:
            # Decode to string
            content_str = file_content.decode("utf-8")
        except UnicodeDecodeError as exc:
            logger.error("Failed to decode MT940 file: %s", exc)
            raise ValueError("Failed to decode MT940 file. Ensure it is UTF-8 encoded.") from exc

        try:
            mt940_parser = mt940.MT940(StringIO(content_str))
            mt940_statements = getattr(mt940_parser, "statements", None) or []
        except Exception as exc:
            logger.error("mt940 library failed to parse content: %s", exc)
            raise ValueError(f"Failed to parse MT940 content with mt940 library: {exc}") from exc

        if not mt940_statements:
            raise ValueError("MT940 file did not contain any statements.")

        transactions: List[BankTransaction] = []
        total_debits = 0.0
        total_credits = 0.0

        # Process only the first statement for now
        stmt = mt940_statements[0]
        start_balance = getattr(stmt, "start_balance", None)
        end_balance = getattr(stmt, "end_balance", None)

        opening_balance = float(start_balance.amount) if start_balance else 0.0
        closing_balance = float(end_balance.amount) if end_balance else 0.0
        currency = (start_balance.currency if start_balance else None) or "EUR"

        from_date = datetime.combine(start_balance.date, datetime.min.time()) if start_balance else datetime.utcnow()
        to_date = datetime.combine(end_balance.date, datetime.min.time()) if end_balance else datetime.utcnow()

        for trans in getattr(stmt, "transactions", []):
            try:
                amount_value = float(trans.amount)
                trans_type = TransactionType.CREDIT if amount_value > 0 else TransactionType.DEBIT
                amount = abs(amount_value)

                if trans_type == TransactionType.DEBIT:
                    total_debits += amount
                else:
                    total_credits += amount

                booking_date = (
                    datetime.combine(trans.booking, datetime.min.time())
                    if getattr(trans, "booking", None)
                    else None
                )
                transaction_date = booking_date or datetime.combine(trans.date, datetime.min.time())

                transaction = BankTransaction(
                    organization_id=self.organization_id,
                    bank_account_id=self.bank_account_id,
                    transaction_date=transaction_date,
                    value_date=booking_date or transaction_date,
                    booking_date=booking_date,
                    transaction_type=trans_type,
                    amount=amount,
                    currency=currency,
                    reference=trans.reference or trans.institution_reference,
                    transaction_id=trans.id,
                    description=trans.description,
                    additional_info=getattr(trans, "additional_data", None),
                    status=TransactionStatus.PENDING,
                    match_status=MatchStatus.UNMATCHED,
                    imported_by=imported_by,
                    raw_data={
                        "reference": trans.reference,
                        "institution_reference": trans.institution_reference,
                        "additional_data": trans.additional_data,
                        "description": trans.description,
                    },
                )

                transactions.append(transaction)

            except Exception as exc:
                logger.error("Error parsing MT940 transaction: %s", exc)
                continue

        if not transactions:
            raise ValueError("No valid transactions were parsed from MT940 file.")

        statement = BankStatement(
            organization_id=self.organization_id,
            bank_account_id=self.bank_account_id,
            format=BankStatementFormat.MT940,
            file_name=file_name,
            file_hash=file_hash,
            statement_number=getattr(stmt, "statement", None),
            statement_date=datetime.utcnow(),
            from_date=from_date,
            to_date=to_date,
            opening_balance=opening_balance,
            closing_balance=closing_balance,
            total_debits=total_debits,
            total_credits=total_credits,
            transaction_count=len(transactions),
            imported_by=imported_by,
            currency=currency,
        )

        return statement, transactions

    def _parse_pdf(
        self,
        file_content: bytes,
        file_name: str,
        file_hash: str,
        imported_by: Optional[str],
    ) -> Tuple[BankStatement, List[BankTransaction]]:
        """
        Parse PDF-based bank statements by extracting tabular data.
        """
        if pdfplumber is None:
            raise ValueError(
                "PDF parsing requires the optional dependency 'pdfplumber'. "
                "Install it (pip install pdfplumber) and try again."
            )

        try:
            raw_rows = self._extract_pdf_rows(file_content)
        except Exception as exc:
            logger.error("Failed to extract tables from PDF: %s", exc)
            raise ValueError(
                "Failed to extract tabular data from the PDF. "
                "Ensure the statement is not password protected and follows a tabular layout."
            ) from exc

        if not raw_rows:
            raise ValueError(
                "No tabular content detected in the PDF statement. "
                "Please upload a PDF export that contains a transaction table or use CSV instead."
            )

        transactions: List[BankTransaction] = []
        opening_balance = None
        closing_balance = None
        total_debits = 0.0
        total_credits = 0.0
        min_date = None
        max_date = None
        detected_currency: Optional[str] = None
        row_errors: List[str] = []

        for row_num, row in enumerate(raw_rows, start=1):
            try:
                sanitized_row = {
                    key: value
                    for key, value in row.items()
                    if key is not None and not str(key).startswith("__")
                }
                normalized_row = self._normalize_row_keys(sanitized_row)

                # Parse dates
                date_value = self._get_first_value(
                    normalized_row,
                    "date",
                    "transaction_date",
                    "booking_date",
                )
                if date_value is None:
                    raise ValueError("Missing transaction date")

                trans_date = self._parse_date(str(date_value))
                value_date = self._parse_date(
                    self._get_first_value(normalized_row, "value_date")
                ) if self._get_first_value(normalized_row, "value_date") else trans_date

                if min_date is None or trans_date < min_date:
                    min_date = trans_date
                if max_date is None or trans_date > max_date:
                    max_date = trans_date

                amount, trans_type = self._extract_amount_and_type(normalized_row)
                if trans_type == TransactionType.DEBIT:
                    total_debits += amount
                else:
                    total_credits += amount

                balance_raw = self._get_first_value(
                    normalized_row,
                    "balance",
                    "balance_after",
                    "running_balance",
                )
                balance_value = self._parse_decimal(balance_raw) if balance_raw else None
                if balance_value is not None:
                    if opening_balance is None:
                        opening_balance = balance_value
                    closing_balance = balance_value

                currency = (
                    self._get_first_value(normalized_row, "currency", "ccy")
                    or detected_currency
                    or "EUR"
                )
                detected_currency = detected_currency or currency

                transaction = BankTransaction(
                    organization_id=self.organization_id,
                    bank_account_id=self.bank_account_id,
                    transaction_date=trans_date,
                    value_date=value_date,
                    booking_date=self._parse_optional_date(
                        self._get_first_value(normalized_row, "booking_date")
                    ),
                    transaction_type=trans_type,
                    amount=amount,
                    currency=currency,
                    reference=self._get_first_value(
                        normalized_row,
                        "reference",
                        "ref",
                        "utr",
                        "utr_no",
                    ),
                    description=self._get_first_value(
                        normalized_row,
                        "description",
                        "details",
                        "narration",
                        "particulars",
                        "remarks",
                    ),
                    counterparty_name=self._get_first_value(
                        normalized_row,
                        "counterparty_name",
                        "beneficiary",
                        "payee",
                        "payer",
                        "customer_name",
                    ),
                    counterparty_account=self._get_first_value(
                        normalized_row,
                        "counterparty_account",
                        "account_number",
                        "account_no",
                        "iban",
                        "iban_no",
                    ),
                    balance_after=balance_value,
                    status=TransactionStatus.PENDING,
                    match_status=MatchStatus.UNMATCHED,
                    imported_by=imported_by,
                    raw_data=self._sanitize_raw_row(sanitized_row),
                )

                transactions.append(transaction)

            except Exception as exc:
                error_message = f"Row {row_num}: {exc}"
                row_errors.append(error_message)
                logger.error("Error parsing PDF row %s: %s", row_num, exc)
                continue

        if not transactions:
            sample_errors = "; ".join(row_errors[:3]) if row_errors else "No parsable rows found."
            raise ValueError(f"No valid transactions were parsed from PDF. {sample_errors}")

        if row_errors:
            logger.warning(
                "PDF parsing skipped %s rows due to validation errors. Sample: %s",
                len(row_errors),
                "; ".join(row_errors[:3]),
            )

        statement = BankStatement(
            organization_id=self.organization_id,
            bank_account_id=self.bank_account_id,
            format=BankStatementFormat.PDF,
            file_name=file_name,
            file_hash=file_hash,
            statement_date=datetime.utcnow(),
            from_date=min_date or datetime.utcnow(),
            to_date=max_date or datetime.utcnow(),
            opening_balance=opening_balance or 0.0,
            closing_balance=closing_balance or 0.0,
            total_debits=total_debits,
            total_credits=total_credits,
            transaction_count=len(transactions),
            imported_by=imported_by,
            currency=detected_currency or "EUR",
        )

        return statement, transactions

    def _extract_pdf_rows(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Extract tabular rows from a PDF statement."""
        if pdfplumber is None:
            return []

        rows: List[Dict[str, Any]] = []
        with pdfplumber.open(BytesIO(file_content)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                if not tables:
                    single_table = page.extract_table()
                    if single_table:
                        tables = [single_table]

                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    header = self._normalize_pdf_header_row(table[0])
                    if not any(header):
                        continue

                    for row in table[1:]:
                        parsed_row = self._build_pdf_row_dict(header, row)
                        if not parsed_row:
                            continue
                        parsed_row["__page__"] = page_index
                        rows.append(parsed_row)

        return rows

    def _normalize_pdf_header_row(self, header_row: List[Any]) -> List[Optional[str]]:
        """Normalize PDF header cells to canonical column names."""
        normalized: List[Optional[str]] = []
        seen: Dict[str, int] = {}
        for cell in header_row:
            col_name = self._normalize_pdf_header_cell(cell)
            if col_name and col_name in seen:
                seen[col_name] += 1
                col_name = f"{col_name}_{seen[col_name]}"
            elif col_name:
                seen[col_name] = 0
            normalized.append(col_name)
        return normalized

    @staticmethod
    def _normalize_pdf_header_cell(cell: Any) -> Optional[str]:
        if cell is None:
            return None

        text = str(cell).strip().lower()
        if not text:
            return None

        text = " ".join(text.replace("\n", " ").split())
        text = re.sub(r"[^\w]+", "_", text)
        text = text.strip("_")

        synonyms = {
            "txn_date": "date",
            "transaction_date": "date",
            "trans_date": "date",
            "transaction_dt": "date",
            "booking_date": "booking_date",
            "posting_date": "booking_date",
            "value_date": "value_date",
            "valuedate": "value_date",
            "value_dt": "value_date",
            "date": "date",
            "narration": "description",
            "details": "description",
            "particulars": "description",
            "remarks": "description",
            "description": "description",
            "reference": "reference",
            "ref": "reference",
            "ref_no": "reference",
            "utr": "reference",
            "utr_no": "reference",
            "instrument_no": "reference",
            "cheque_no": "reference",
            "amount": "amount",
            "amt": "amount",
            "transaction_amount": "amount",
            "debit": "debit",
            "dr": "debit",
            "withdrawal": "debit",
            "debit_amount": "debit",
            "amount_dr": "debit",
            "credit": "credit",
            "cr": "credit",
            "deposit": "credit",
            "credit_amount": "credit",
            "amount_cr": "credit",
            "balance": "balance",
            "balance_amount": "balance",
            "balance_amt": "balance",
            "running_balance": "balance",
            "closing_balance": "balance",
            "opening_balance": "balance",
            "currency": "currency",
            "currency_code": "currency",
            "ccy": "currency",
            "beneficiary": "counterparty_name",
            "counterparty": "counterparty_name",
            "payer": "counterparty_name",
            "payee": "counterparty_name",
            "customer_name": "counterparty_name",
            "account_number": "counterparty_account",
            "account_no": "counterparty_account",
            "account": "counterparty_account",
            "iban": "counterparty_account",
            "iban_no": "counterparty_account",
        }

        if text in synonyms:
            return synonyms[text]

        # Fuzzy matching for headers that embed currency codes or extra words
        if "date" in text and any(token in text for token in ("transaction", "txn", "booking")):
            return "date" if "value" not in text else "value_date"
        if "value" in text and "date" in text:
            return "value_date"
        if "debit" in text or text.endswith("_dr"):
            return "debit"
        if "credit" in text or text.endswith("_cr"):
            return "credit"
        if "amount" in text or "amt" in text:
            return "amount"
        if "balance" in text:
            return "balance"
        if "currency" in text or "ccy" in text:
            return "currency"
        if "reference" in text or "ref" in text or "utr" in text:
            return "reference"

        return text

    @staticmethod
    def _build_pdf_row_dict(
        header: List[Optional[str]],
        row: List[Any],
    ) -> Dict[str, Any]:
        parsed: Dict[str, Any] = {}
        has_value = False

        for col_name, value in zip(header, row):
            if col_name is None:
                continue
            if value is None:
                continue

            if isinstance(value, str):
                cleaned = value.strip()
                if not cleaned:
                    continue
                parsed[col_name] = cleaned
            else:
                parsed[col_name] = value

            if parsed[col_name] not in ("", None):
                has_value = True

        return parsed if has_value else {}

    @staticmethod
    def _parse_date(date_str: Optional[str]) -> datetime:
        """
        Parse date string to datetime

        Supports common formats:
        - YYYY-MM-DD
        - DD/MM/YYYY
        - DD.MM.YYYY
        - YYYYMMDD
        """
        if not date_str:
            return datetime.utcnow()

        # Try common formats
        formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%Y",
            "%Y%m%d",
            "%d-%m-%Y",
            "%Y/%m/%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue

        # If all fail, return current date
        logger.warning(f"Could not parse date: {date_str}")
        return datetime.utcnow()

    def _parse_optional_date(self, date_value: Optional[Any]) -> Optional[datetime]:
        """Parse optional date fields safely."""
        if date_value is None:
            return None
        try:
            return self._parse_date(str(date_value))
        except Exception:
            return None

    @staticmethod
    def _sanitize_raw_row(row: Dict[Any, Any]) -> Dict[str, Any]:
        """Ensure raw_data keys are strings for Pydantic validation."""
        sanitized: Dict[str, Any] = {}
        for key, value in row.items():
            if key is None:
                continue
            sanitized[str(key)] = value
        return sanitized

    @staticmethod
    def _parse_decimal(value: Any) -> Optional[float]:
        """Parse numeric strings that may contain locale or currency characters."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None

            sign = 1
            if cleaned.startswith("(") and cleaned.endswith(")"):
                sign = -1
                cleaned = cleaned[1:-1]

            cleaned = (
                cleaned.replace(",", "")
                .replace(" ", "")
                .replace("€", "")
                .replace("$", "")
                .replace("+", "")
            )

            suffixes = {"CR": 1, "DR": -1}
            for suffix, suffix_sign in suffixes.items():
                if cleaned.upper().endswith(suffix):
                    sign *= suffix_sign
                    cleaned = cleaned[:-len(suffix)]
                    break

            try:
                return float(Decimal(cleaned)) * sign
            except (InvalidOperation, ValueError):
                return None

        return None

    def _extract_amount_and_type(self, row: Dict[str, Any]) -> Tuple[float, TransactionType]:
        """
        Determine transaction amount and type from flexible CSV schemas.
        `row` must have lowercase, trimmed keys (use `_normalize_row_keys`).
        Raises ValueError if the amount cannot be determined or is zero.
        """
        amount_fields = [
            "amount",
            "transaction_amount",
            "value",
            "amt",
        ]
        paired_fields = [
            ("credit", "debit"),
            ("credit_amount", "debit_amount"),
            ("paid_in", "paid_out"),
            ("deposit", "withdrawal"),
            ("money_in", "money_out"),
        ]
        indicator_fields = [
            "transaction_type",
            "type",
            "credit_debit_indicator",
            "credit_debit",
            "cr_dr",
            "dr_cr",
            "direction",
            "indicator",
            "debitcredit",
        ]
        indicator_type = None
        for indicator_field in indicator_fields:
            indicator_type = self._infer_type_from_indicator(row.get(indicator_field))
            if indicator_type:
                break

        for field in amount_fields:
            parsed = self._parse_decimal(row.get(field))
            if parsed is None or parsed == 0:
                continue
            if parsed > 0:
                return parsed, TransactionType.CREDIT
            return abs(parsed), TransactionType.DEBIT

        for credit_field, debit_field in paired_fields:
            credit_value = self._parse_decimal(row.get(credit_field))
            if credit_value and credit_value > 0:
                return credit_value, TransactionType.CREDIT

            debit_value = self._parse_decimal(row.get(debit_field))
            if debit_value and debit_value > 0:
                return debit_value, TransactionType.DEBIT

        # Fallback: scan all columns containing "amount"/"amt"
        for key, value in row.items():
            if "amount" not in key and "amt" not in key:
                continue
            if any(exclusion in key for exclusion in ("opening_balance", "closing_balance", "balance")):
                continue

            parsed = self._parse_decimal(value)
            if parsed is None or parsed == 0:
                continue

            # Infer type from column name if possible
            lower_key = key.lower()
            if "credit" in lower_key or lower_key.endswith("cr"):
                return abs(parsed), TransactionType.CREDIT
            if "debit" in lower_key or lower_key.endswith("dr"):
                return abs(parsed), TransactionType.DEBIT

            if indicator_type:
                return abs(parsed), indicator_type

            # Default to credit for positive, debit for negative even though we excluded 0
            if parsed > 0:
                return parsed, TransactionType.CREDIT
            return abs(parsed), TransactionType.DEBIT

        available_fields = ", ".join(row.keys())
        raise ValueError(
            "Unable to determine transaction amount (missing amount/credit/debit values). "
            f"Available columns: {available_fields}"
        )

    @staticmethod
    def _normalize_row_keys(row: Dict[Any, Any]) -> Dict[str, Any]:
        """Return a new dict with lowercase, trimmed keys for case-insensitive lookups."""
        normalized: Dict[str, Any] = {}
        for key, value in row.items():
            if key is None:
                continue
            normalized[str(key).strip().lower()] = value
        return normalized

    @staticmethod
    def _get_first_value(row: Dict[str, Any], *keys: str) -> Optional[Any]:
        """Fetch the first available key (case-insensitive) from normalized row."""
        for key in keys:
            if key is None:
                continue
            lookup = key.strip().lower()
            if lookup in row:
                return row[lookup]
        return None

    @staticmethod
    def _infer_type_from_indicator(indicator: Optional[Any]) -> Optional[TransactionType]:
        if indicator is None:
            return None
        text = str(indicator).strip().lower()
        if not text:
            return None

        credit_tokens = {"credit", "cr", "c", "in", "+", "deposit"}
        debit_tokens = {"debit", "dr", "d", "out", "-", "withdrawal"}

        if text in credit_tokens:
            return TransactionType.CREDIT
        if text in debit_tokens:
            return TransactionType.DEBIT

        if "credit" in text:
            return TransactionType.CREDIT
        if "debit" in text:
            return TransactionType.DEBIT

        return None

    @staticmethod
    def _strip_leading_metadata_lines(content_str: str) -> str:
        """
        Remove introductory metadata lines so the CSV header starts with actual column names.
        Detects the first line that contains a delimiter and any expected column keyword.
        """
        lines = content_str.splitlines()
        header_keywords = [
            "date",
            "transaction",
            "amount",
            "debit",
            "credit",
            "balance",
            "description",
            "details",
            "reference",
        ]
        delimiters = [",", ";", "\t", "|"]

        def looks_like_header(line: str) -> bool:
            lower = line.lower()
            if not any(delim in line for delim in delimiters):
                return False
            return any(keyword in lower for keyword in header_keywords)

        for idx, line in enumerate(lines):
            if looks_like_header(line):
                stripped = "\n".join(lines[idx:])
                if idx > 0:
                    logger.info("Skipped %s metadata lines before CSV header detection", idx)
                return stripped

        return content_str

    def _get_csv_string(self, file_content: bytes, file_name: str) -> str:
        """Return text content for CSV parsing, converting Excel if necessary."""
        if self._is_excel_content(file_content, file_name):
            return self._excel_to_csv_string(file_content)

        encodings = [
            "utf-8-sig",
            "utf-8",
            "latin-1",
            "iso-8859-1",
            "cp1252",
            "windows-1252",
        ]

        for encoding in encodings:
            try:
                decoded = file_content.decode(encoding)
                logger.info("Successfully decoded CSV with %s encoding", encoding)
                return decoded
            except (UnicodeDecodeError, AttributeError):
                continue

        raise ValueError("Could not decode CSV file with any supported encoding")

    def _is_excel_content(self, file_content: bytes, file_name: Optional[str]) -> bool:
        """Detect if the uploaded file is actually an Excel workbook."""
        excel_extensions = (".xlsx", ".xlsm", ".xlsb", ".xls")
        if file_name and file_name.lower().endswith(excel_extensions):
            return True

        # Excel OpenXML files are ZIP archives that start with PK header
        try:
            if not zipfile.is_zipfile(BytesIO(file_content)):
                return False
            with zipfile.ZipFile(BytesIO(file_content)) as archive:
                names = archive.namelist()
                return "[Content_Types].xml" in names or any(
                    name.endswith((".xml", ".rels")) for name in names
                )
        except Exception:
            return False

    def _excel_to_csv_string(self, file_content: bytes) -> str:
        """Convert first worksheet of an Excel file to CSV string."""
        if load_workbook is None:
            raise ValueError(
                "Excel file detected but openpyxl is not installed. "
                "Install openpyxl or upload CSV exports instead."
            )

        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to read Excel file: {exc}") from exc
        worksheet = workbook.active

        output = StringIO()
        csv_writer = csv.writer(output)
        rows_written = 0

        for row in worksheet.iter_rows(values_only=True):
            sanitized_row = ["" if cell is None else cell for cell in row]
            csv_writer.writerow(sanitized_row)
            rows_written += 1

        if rows_written == 0:
            raise ValueError("Excel file does not contain any rows to import")

        return output.getvalue()

    @staticmethod
    def detect_format(file_content: bytes, file_name: str) -> BankStatementFormat:
        """
        Auto-detect bank statement format based on content

        Args:
            file_content: Raw file bytes
            file_name: Original filename

        Returns:
            Detected BankStatementFormat
        """
        # Check file extension
        file_name_lower = file_name.lower()

        if file_name_lower.endswith(".xml"):
            # Try to detect if it's CAMT.053
            try:
                content_str = file_content.decode("utf-8")
                if "camt.053" in content_str or "BkToCstmrStmt" in content_str:
                    return BankStatementFormat.CAMT053
            except:
                pass

        if file_name_lower.endswith(".csv"):
            return BankStatementFormat.CSV

        if file_name_lower.endswith((".txt", ".mt940", ".sta")):
            return BankStatementFormat.MT940

        if file_name_lower.endswith(".pdf"):
            return BankStatementFormat.PDF

        # Content-based detection
        try:
            content_str = file_content.decode("utf-8")

            # Check for CAMT.053 XML structure
            if "<?xml" in content_str and "camt.053" in content_str:
                return BankStatementFormat.CAMT053

            # Check for MT940 structure
            if ":20:" in content_str and ":25:" in content_str:
                return BankStatementFormat.MT940

            # Default to CSV
            return BankStatementFormat.CSV

        except:
            if file_content.startswith(b"%PDF"):
                return BankStatementFormat.PDF
            # If all fails, assume CSV
            return BankStatementFormat.CSV

    @staticmethod
    def _looks_like_pdf(file_content: bytes, file_name: Optional[str]) -> bool:
        if file_name and file_name.lower().endswith(".pdf"):
            return True
        return file_content.startswith(b"%PDF")
